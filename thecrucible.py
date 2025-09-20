import argparse
import sqlite3  # SQLite for the database
import csv
from openpyxl import Workbook  # Excel manipulation
import subprocess
import os
import re
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage




# Command-line arguments
parser = argparse.ArgumentParser(description="Process Baselight and Xytech files.")
parser.add_argument("--baselight", type=str, help="Path to Baselight file")
parser.add_argument("--xytech", type=str, help="Path to Xytech file")
parser.add_argument("--process", type=str, help="Path to the video file for processing")
parser.add_argument("--outputXLS", type=str, help="Path to output XLS file")
parser.add_argument("--outputCSV", type=str, help="Path to output CSV file")
args = parser.parse_args()


# Functions from Proj1

def read_file(file_path):
    with open(file_path, 'r') as file:
        return file.readlines()


def parse_baselight(data):
    parsed_frames = []
    for line in data:
        if "<err>" in line:
            continue  # Skip lines with errors
        components = line.strip().split()
        if len(components) < 2:
            continue  # Skip invalid lines

        filename = components[0]
        frame_data = components[1:]
        cleaned_ranges = []

        # Combine frame ranges into a unified string
        for frame in frame_data:
            cleaned_ranges.append(frame)

        # Join and clean ranges
        full_range = ', '.join(cleaned_ranges)
        parsed_frames.append((filename, full_range))
    return parsed_frames



def validate_numeric(value):
    """Ensure value is numeric or extract numeric part of a string."""
    match = re.match(r"^\d+", value)  # Extract numeric prefix
    if match:
        return int(match.group(0))
    return None

def validate_frame_ranges(video_length, frame_ranges):
    """
    Validates frame ranges against the video duration.
    Returns valid and invalid frame ranges.
    """
    valid_frames = []
    invalid_frames = []
    fps = 24  # Assuming 24 frames per second
    video_frames = int(video_length * fps)

    for filename, frames in frame_ranges:
        try:
            start, end = map(int, frames.split('-'))
            if start <= video_frames and end <= video_frames:
                valid_frames.append((filename, frames))
            else:
                invalid_frames.append((filename, frames))
        except ValueError:
            invalid_frames.append((filename, frames))

    return valid_frames, invalid_frames



def format_frames(frames):
    """Format frames with ranges and individual numbers."""
    ranges = []
    start = frames[0]
    end = frames[0]

    for i in range(1, len(frames)):
        if frames[i] == end + 1:
            end = frames[i]
        else:
            if start == end:
                ranges.append(f"{start}")
            else:
                ranges.append(f"{start}-{end}")
            start = frames[i]
            end = frames[i]

    # Add the last range or single frame
    if start == end:
        ranges.append(f"{start}")
    else:
        ranges.append(f"{start}-{end}")

    return ", ".join(ranges)


def parse_xytech(data):
    parsed_orders = []
    for line in data:
        if '/' in line:
            components = line.strip().split('/')
            producer = components[0].strip()
            operator = components[1].strip()
            order_info = components[-1].strip()
            parsed_orders.append((producer, operator, order_info))
    return parsed_orders


def clean_data(data):
    cleaned_data = []
    for item in data:
        cleaned_item = tuple(
            elem.replace('<null>', '').strip()
            if isinstance(elem, str) else elem
            for elem in item
        )
        cleaned_data.append(cleaned_item)
    return cleaned_data


def match_data(baselight, xytech):
    matched_data = []
    for i in range(min(len(baselight), len(xytech))):
        producer, operator, order_info = xytech[i]
        filename, frames = baselight[i]
        matched_data.append((producer, operator, order_info, filename, frames))
    return matched_data

def export_to_xls_combined_with_images_and_timestamps(data, output_path, video_path):
    """
    Export matched data, including embedded thumbnail images, frame ranges, and corresponding timestamps.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Data"
    
    # Add headers
    ws.append(["Producer", "Operator", "Order Info", "Filename", "Frame Range", "Timestamp", "Thumbnail"])
    
    # Set column widths for better display
    ws.column_dimensions['A'].width = 20  # Producer
    ws.column_dimensions['B'].width = 20  # Operator
    ws.column_dimensions['C'].width = 30  # Order Info
    ws.column_dimensions['D'].width = 50  # Filename
    ws.column_dimensions['E'].width = 20  # Frame Range
    ws.column_dimensions['F'].width = 15  # Timestamp
    ws.column_dimensions['G'].width = 15  # Thumbnail

    # Populate rows with data and images
    for i, row in enumerate(data, start=2):  # Start from the second row
        producer, operator, order_info, filename, frames, thumbnail = row
        try:
            # Calculate timestamp
            start, end = map(int, frames.split('-'))
            middle_frame = (start + end) // 2
            timestamp = middle_frame / 24  # Assuming 24 fps
            
            # Add text data to the row
            ws.append([producer, operator, order_info, filename, frames, f"{timestamp:.2f} sec"])
            
            # Embed the image in the thumbnail column
            if thumbnail and os.path.exists(thumbnail):
                img = ExcelImage(thumbnail)
                img.width, img.height = 96, 74
                img_anchor = f"G{i}"  # Ensure the cell corresponds to the correct row
                ws.add_image(img, img_anchor)
                print(f"Embedded thumbnail {thumbnail} at {img_anchor}")
            else:
                ws[f"G{i}"] = "No Thumbnail"
        except Exception as e:
            print(f"Error processing frame range {frames}: {e}")
            ws.append([producer, operator, order_info, filename, frames, "Error", "No Thumbnail"])


    # Save the workbook
    wb.save(output_path)
    print(f"Combined data with thumbnails and timestamps exported to XLS: {output_path}")



def create_thumbnail(video_path, frame_range, output_dir):
    """
    Creates a thumbnail for the middle frame of a given frame range.
    Returns the path to the thumbnail.
    """
    start, end = map(int, frame_range.split('-'))
    middle_frame = (start + end) // 2
    timestamp = middle_frame / 24  # Convert frame to seconds assuming 24 FPS

    # Create the output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Define the output thumbnail image path
    output_image = f"{output_dir}/thumb_{middle_frame}.jpg"

    # Extract the middle frame using ffmpeg
    result = subprocess.run(
        [
            "ffmpeg",
            "-i", video_path,
            "-ss", str(timestamp),
            "-vframes", "1",
            output_image
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE
    )

    # Resize the thumbnail if extraction succeeded
    if result.returncode == 0 and os.path.exists(output_image):
        img = Image.open(output_image)
        img.thumbnail((96, 74))
        img.save(output_image)
        return output_image
    else:
        return None


def populate_database(baselight_data, xytech_data):
    conn = sqlite3.connect("thecrucible_database.db")
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Baselight (
            Filename TEXT,
            Frames TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Xytech (
            Producer TEXT,
            Operator TEXT,
            OrderInfo TEXT
        )
    """)

    cursor.executemany("INSERT INTO Baselight (Filename, Frames) VALUES (?, ?)", baselight_data)
    cursor.executemany("INSERT INTO Xytech (Producer, Operator, OrderInfo) VALUES (?, ?, ?)", xytech_data)

    conn.commit()
    conn.close()

# Add this after parsing the data:
def export_to_xls(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Data"
    ws.append(["Producer", "Operator", "Order Info", "Filename", "Frames", "Timecode Range", "Thumbnail"])
    for row in data:
        producer, operator, order_info, filename, frames = row
        timecode_range = "00:00:00-00:00:05"  # Placeholder logic
        thumbnail_path = f"thumbnails/thumb_{frames.split('-')[0]}.jpg"
        ws.append([producer, operator, order_info, filename, frames, timecode_range, thumbnail_path])
    wb.save(output_path)




def get_video_length(video_path):
    result = subprocess.run(
        [
            "ffprobe",
            "-i", video_path,
            "-v", "error",
            "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1"
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE
    )
    try:
        return float(result.stdout.strip())
    except ValueError:
        print("Error extracting video length.")
        return 0


def export_to_xls(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Data"
    ws.append(["Producer", "Operator", "Order Info", "Filename", "Frames"])
    for row in data:
        ws.append(row)
    wb.save(output_path)

def find_unused_frames(baselight_data, matched_data):
    used_frames = set(row[1] for row in matched_data)  # Extract frames from matched data
    unused_frames = [row for row in baselight_data if row[1] not in used_frames]
    return unused_frames
def export_unused_frames_to_csv(unused_frames, output_path):
    with open(output_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Filename", "Frames"])
        writer.writerows(unused_frames)
    print(f"Unused frames exported to: {output_path}")

def frame_to_timecode(frame, fps=24):
    hours = frame // (3600 * fps)
    minutes = (frame % (3600 * fps)) // (60 * fps)
    seconds = (frame % (60 * fps)) // fps
    frames = frame % fps
    return f"{hours:02}:{minutes:02}:{seconds:02}:{frames:02}"


def render_shot(video_path, frame_range, output_dir):
    start, end = map(int, frame_range.split('-'))
    timestamp_start = start / 24
    duration = (end - start) / 24
    output_file = os.path.join(output_dir, f"shot_{start}_{end}.mp4")
    subprocess.run(
        ["ffmpeg", "-i", video_path, "-ss", str(timestamp_start), "-t", str(duration), output_file],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE
    )
    return output_file


    






def find_matching_ranges(video_length, frame_ranges):
    matching_ranges = []
    fps = 24  # Assume 24 frames per second
    video_frames = int(video_length * fps)  # Convert video length to total frames

    for frame_range in frame_ranges:
        try:
            # Split by commas for mixed ranges and single frames
            parts = frame_range.split(',')
            for part in parts:
                part = part.strip()
                if '-' in part:  # Handle ranges like "1-10"
                    start, end = map(int, part.split('-'))
                else:  # Handle single frames like "15"
                    start = end = int(part)

                # Validate against video frames
                if start <= video_frames and end <= video_frames:
                    matching_ranges.append((start, end))
        except ValueError as e:
            print(f"Skipping invalid frame range: {frame_range} - {e}")

    return matching_ranges

def export_unused_to_csv(unused_data, output_csv):
    with open(output_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Filename", "Unused Frame Range"])
        writer.writerows(unused_data)
    print(f"Unused frames exported to {output_csv}")



# Add this after calculating video length:
thumbnail_dir = "thumbnails"
os.makedirs(thumbnail_dir, exist_ok=True)  # Creates the directory if it doesn't exist

# Main logic
baselight_data = read_file(args.baselight)
xytech_data = read_file(args.xytech)

parsed_baselight = clean_data(parse_baselight(baselight_data))
parsed_xytech = clean_data(parse_xytech(xytech_data))
populate_database(parsed_baselight, parsed_xytech)

video_length = get_video_length(args.process)
frame_ranges = [row[1] for row in parsed_baselight]
matching_ranges = find_matching_ranges(video_length, frame_ranges)

print("Parsed Baselight Data:", parsed_baselight)
print("Parsed Xytech Data:", parsed_xytech)


valid_frames = [(row[0], f"{start}-{end}") for row in parsed_baselight for start, end in matching_ranges]
unused_frames = [(row[0], row[1]) for row in parsed_baselight if row[1] not in [f"{start}-{end}" for start, end in matching_ranges]]

export_unused_to_csv(unused_frames, args.outputCSV)

# Process valid frames and export
matched_data = match_data(valid_frames, parsed_xytech)
for i, row in enumerate(matched_data):
    filename, frames = row[3], row[4]
    try:
        thumbnail = create_thumbnail(args.process, frames, "thumbnails")
        matched_data[i] = (*row, thumbnail)  # Append thumbnail
    except Exception as e:
        print(f"Error creating thumbnail for {filename}: {e}")
        matched_data[i] = (*row, None)  # Append None for failed thumbnails

export_to_xls_combined_with_images_and_timestamps(matched_data, args.outputXLS, args.process)